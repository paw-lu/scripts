"""Rough script to export DataFrame into styled Excel sheet."""
from itertools import product

import pandas as pd

# Update openpyxl to latest version to get this to work, or else it will mess up on timestamps
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, colors
from openpyxl.utils.dataframe import dataframe_to_rows

show_index = False

df = pd.DataFrame(
    {
        "name": [
            "cabbage",
            "cheese",
            "ham",
            "ram",
            "sam",
            "a mini van",
            "chocolate shake",
            "guava juice",
        ],
        "cost": ["2.0", "5.0", "9.0", "1000", "40", "20", "50", "10"],
        "weight": [3, 4, 5, 1, 3, 6, 2, 3],
    }
)

file = "my_style.xlsx"

wb = Workbook()
# pick the active sheet
worksheet = wb.active
worksheet.sheet_view.showGridLines = False  # Turn off gridlines
# Write into the first sheet
for i, r in enumerate(dataframe_to_rows(df, index=show_index, header=True)):
    if show_index and i == 1:
        continue
    worksheet.append(r)


col_iterations = (df.shape[1] + show_index) // 26
xl_col_names = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")

i = 0
while i < col_iterations:
    xl_col_names.extend(
        [i[0] + i[1] for i in product(xl_col_names, "ABCDEFGHIJKLMNOPQRSTUVWXYZ")]
    )
    i += 1

xl_cols = []
for col in range(df.shape[1] + show_index):
    xl_cols.append(xl_col_names[col])

header_font = Font(name="Lucida Console", bold=True, size=12, color="00212121")
value_font = Font(name="Segoe UI", size=12, color="00212121")
index_font = Font(name="Segoe UI", bold=True, size=12, color="00212121")
index_align = Alignment(horizontal="left")
num_rows = df.shape[0]

# get everything in row 1
for cell in worksheet["1"]:
    cell.font = header_font

rectangle = product(xl_cols, range(2, num_rows + 2))
for column, row in rectangle:
    worksheet[f"{column}{row}"].font = value_font

if show_index:
    for cell in worksheet["A"]:
        cell.font = index_font
        cell.alignment = index_align

# note that the "#" in the hex code is replaced with two zeros
alt_fill = PatternFill(fill_type="solid", fgColor="00f5f5f5")

for col in xl_cols:
    for cell in worksheet[col]:
        if cell.row > 1 and cell.row % 2 == 0:
            cell.fill = alt_fill

custom_side = Side(border_style="thin", color="00dadada")
for cell in worksheet["1"]:
    cell.border = Border(bottom=custom_side)

# last_row = df.shape[0] + 1

# custom_side = Side(border_style='hair', color="00548444")

# # define all four sides of the box. Top:
# worksheet['A2'].border = worksheet['A2'].border + Border(top=custom_side)

# # left and right:
# for cell in worksheet['A']:
#     if cell.row >= 2 and cell.row <= last_row:
#         cell.border = cell.border + Border(left=custom_side, right=custom_side)

# # bottom:
# worksheet[f"A{last_row}"].border = worksheet[f"A{last_row}"].border + Border(
#     bottom=custom_side
# )

dims = {}
for row in worksheet.rows:
    for cell in row:
        if cell.value:
            dims[cell.coordinate[0]] = max(
                (dims.get(cell.coordinate[0], 0), len(str(cell.value)))
            )

for col, value in dims.items():
    if show_index and col == "A":
        worksheet.column_dimensions[col].width = 2 * (value + 5)
    else:
        worksheet.column_dimensions[col].width = value + 5

wb.save(filename=file)
wb.close()